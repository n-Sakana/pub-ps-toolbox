#!/usr/bin/env python3
"""Live E2E test for the ISA/MOJ crawler prototype.

This intentionally accesses https://www.moj.go.jp/isa/. It verifies the workbook
against the current live HTML, not just that a file was created.
"""

from __future__ import annotations

import re
import subprocess
import sys
import tempfile
from pathlib import Path
from urllib.parse import urldefrag, urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "crawler.py"
URL_FILE = ROOT / "URL.txt"
CONFIG = ROOT / "config.json"
EXPECTED_MIN_PAGES = 10
EXPECTED_MIN_PDFS = 1
REQUEST_HEADERS = {
    "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36",
    "Accept-Language": "ja,en-US;q=0.9,en;q=0.8",
}


def normalize_space(text: object) -> str:
    return re.sub(r"\s+", " ", str(text or "").replace("\ufeff", "").replace("\u200b", "")).strip()


def fetch_html(url: str) -> str:
    response = requests.get(url, headers=REQUEST_HEADERS, timeout=30)
    response.raise_for_status()
    return response.content.decode("utf-8", errors="replace")


def sheet_records(workbook, sheet_name: str) -> list[dict[str, object]]:
    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value) for value in next(rows)]
    return [dict(zip(headers, row)) for row in rows]


def source_pdf_links(url: str) -> set[str]:
    soup = BeautifulSoup(fetch_html(url), "html.parser")
    links = set()
    for a in soup.find_all("a", href=True):
        linked, _fragment = urldefrag(urljoin(url, a["href"]))
        if ".pdf" in linked.lower():
            links.add(linked)
    return links


def main() -> int:
    with tempfile.TemporaryDirectory(prefix="moj-isa-crawler-e2e-") as tmpdir:
        tmp = Path(tmpdir)
        output = tmp / "moj_isa_crawl.xlsx"
        download_dir = tmp / "pdfs"
        graph_dir = tmp / "graphs"
        log_file = tmp / "crawler.log"
        error_log_file = tmp / "errors.txt"
        cmd = [
            sys.executable,
            str(SCRIPT),
            "--url-file",
            str(URL_FILE),
            "--config",
            str(CONFIG),
            "--output",
            str(output),
            "--download-dir",
            str(download_dir),
            "--graph-dir",
            str(graph_dir),
            "--log-file",
            str(log_file),
            "--error-log-file",
            str(error_log_file),
            "--progress-every-pages",
            "1",
            "--progress-every-pdfs",
            "50",
            "--sleep",
            "0",
            "--max-pages",
            "15",
            "--download-pdfs",
            "--max-pdf-downloads",
            "1",
        ]
        completed = subprocess.run(cmd, cwd=ROOT, text=True, capture_output=True)
        print(completed.stdout, end="")
        if completed.stderr:
            print(completed.stderr, file=sys.stderr, end="")
        if completed.returncode != 0:
            return completed.returncode
        if not output.exists() or output.stat().st_size == 0:
            raise AssertionError("Workbook was not created")
        if not log_file.exists() or log_file.stat().st_size == 0:
            raise AssertionError("Run log was not created")
        if not error_log_file.exists() or error_log_file.stat().st_size == 0:
            raise AssertionError("Error log was not created")
        log_text = log_file.read_text(encoding="utf-8")
        if "PAGE_PROGRESS" not in log_text or "PDF_DOWNLOAD_OK" not in log_text or "PDF_FINAL_VERIFY" not in log_text or "GRAPH_PHASE_DONE" not in log_text or "DONE" not in log_text:
            raise AssertionError("Run log does not contain expected progress markers")
        error_text = error_log_file.read_text(encoding="utf-8")
        if "errors: 0" not in error_text or "No errors." not in error_text:
            raise AssertionError("Empty error report was not written")

        workbook = load_workbook(output, read_only=True, data_only=True)
        for sheet in ["Pages", "PDFs", "Links", "Errors", "Summary", "Stats", "DepthStats", "SectionStats", "PdfStats", "ErrorStats", "TopPages", "Graphs"]:
            if sheet not in workbook.sheetnames:
                raise AssertionError(f"Missing sheet: {sheet}")
        graph_files = sorted(graph_dir.glob("*.png"))
        if len(graph_files) < 3:
            raise AssertionError(f"Expected graph PNGs, got {len(graph_files)} in {graph_dir}")
        dot_files = sorted(graph_dir.glob("*.dot"))
        if len(dot_files) < 2:
            raise AssertionError(f"Expected Graphviz DOT files, got {len(dot_files)} in {graph_dir}")
        status_file = graph_dir / "graphviz_status.txt"
        if not status_file.exists() or "Graphviz status" not in status_file.read_text(encoding="utf-8"):
            raise AssertionError("Graphviz status file was not written")

        pages = sheet_records(workbook, "Pages")
        pdfs = sheet_records(workbook, "PDFs")
        links = sheet_records(workbook, "Links")
        if len(pages) < EXPECTED_MIN_PAGES:
            raise AssertionError(f"Expected at least {EXPECTED_MIN_PAGES} pages, got {len(pages)}")
        if len(pdfs) < EXPECTED_MIN_PDFS:
            raise AssertionError(f"Expected at least {EXPECTED_MIN_PDFS} PDF refs, got {len(pdfs)}")
        if not links:
            raise AssertionError("Links sheet is empty")

        # Page rows are checked against live HTML titles/text.
        for row in pages[:5]:
            url = normalize_space(row["url"])
            live_soup = BeautifulSoup(fetch_html(url), "html.parser")
            live_text = normalize_space(live_soup.get_text(" ", strip=True))
            title = normalize_space(row["title"])
            if title and title[:20] not in live_text:
                raise AssertionError(f"Page title probe not found in live HTML: {url}: {title!r}")
            body = normalize_space(row["body_text"])
            if body:
                probe = body[:30]
                if probe not in live_text:
                    raise AssertionError(f"Body text probe not found in live HTML: {url}: {probe!r}")

        # PDF rows are checked against the live source page's actual hrefs.
        by_source: dict[str, set[str]] = {}
        for row in pdfs:
            source = normalize_space(row["source_page_url"])
            pdf_url = normalize_space(row["pdf_url"])
            if source not in by_source:
                by_source[source] = source_pdf_links(source)
            if pdf_url not in by_source[source]:
                raise AssertionError(f"PDF URL is not present on live source page: {pdf_url}")

        downloaded = [row for row in pdfs if str(row.get("downloaded", "")).lower() == "true"]
        if not downloaded:
            raise AssertionError("No PDF row was marked downloaded")
        downloaded_path = Path(normalize_space(downloaded[0]["saved_path"]))
        sha = normalize_space(downloaded[0]["sha256"])
        if not downloaded_path.exists() or downloaded_path.stat().st_size == 0:
            raise AssertionError(f"Downloaded PDF missing or empty: {downloaded_path}")
        if len(sha) != 64:
            raise AssertionError(f"Invalid sha256: {sha}")
        if normalize_space(downloaded[0].get("first_bytes_hex", "")) != "255044462d":
            raise AssertionError("Downloaded PDF first bytes were not recorded as %PDF-")
        if str(downloaded[0].get("post_write_readback_ok", "")).lower() != "true":
            raise AssertionError("Downloaded PDF post-write readback check did not pass")

        print(f"E2E OK: pages={len(pages)}, pdf_refs={len(pdfs)}, downloaded={downloaded_path.name}, graphs={len(graph_files)}, dot={len(dot_files)}, log={log_file.name}, errors={error_log_file.name}")
        return 0


if __name__ == "__main__":
    raise SystemExit(main())
