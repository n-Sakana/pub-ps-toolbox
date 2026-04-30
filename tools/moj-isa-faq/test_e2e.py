#!/usr/bin/env python3
"""Live E2E test for tools/moj-isa-faq/qa_scraper.py.

This test intentionally accesses the official ISA/MOJ website. It does not only
check that an Excel file exists; it also re-fetches the source pages and verifies
that the workbook contents match visible text on the live HTML pages.
"""

from __future__ import annotations

import re
import subprocess
import sys
import tempfile
from pathlib import Path
from urllib.parse import urldefrag, urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
SCRIPT = ROOT / "qa_scraper.py"
URL_FILE = ROOT / "URL.txt"
EXPECTED_PAGES = 8
MIN_EXPECTED_ROWS = 400
REQUEST_HEADERS = {
    "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/123.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "ja,en-US;q=0.9,en;q=0.8",
}


def normalize_space(text: object) -> str:
    return re.sub(r"\s+", " ", str(text or "").replace("\ufeff", "").replace("\u200b", "")).strip()


def fetch_html(url: str) -> str:
    response = requests.get(url, headers=REQUEST_HEADERS, timeout=30)
    response.raise_for_status()
    return response.content.decode("utf-8", errors="replace")


def read_index_url() -> str:
    urls = [
        line.strip()
        for line in URL_FILE.read_text(encoding="utf-8-sig").splitlines()
        if line.strip() and not line.lstrip().startswith("#")
    ]
    if len(urls) != 1:
        raise AssertionError(f"URL.txt must contain exactly one URL, got {len(urls)}")
    return urls[0]


def discover_index_pages(index_url: str) -> list[tuple[str, str]]:
    soup = BeautifulSoup(fetch_html(index_url), "html.parser")
    root = soup.select_one("#contentsArea") or soup.select_one("#tex") or soup
    pages: list[tuple[str, str]] = []
    seen: set[str] = set()
    for a in root.select("ul.menuList01 a[href]"):
        title = normalize_space(a.get_text(" ", strip=True))
        url, _fragment = urldefrag(urljoin(index_url, a.get("href", "")))
        if title and url not in seen:
            seen.add(url)
            pages.append((title, url))
    return pages


def sheet_records(workbook, sheet_name: str) -> list[dict[str, object]]:
    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value) for value in next(rows)]
    return [dict(zip(headers, row)) for row in rows]


def visible_text_cache(urls: set[str]) -> dict[str, str]:
    cache: dict[str, str] = {}
    for url in sorted(urls):
        base_url, _fragment = urldefrag(url)
        if base_url in cache:
            continue
        soup = BeautifulSoup(fetch_html(base_url), "html.parser")
        cache[base_url] = normalize_space(soup.get_text(" ", strip=True))
    return cache


def assert_text_appears(label: str, expected: object, source_text: str, row_no: int) -> None:
    text = normalize_space(expected)
    if not text:
        raise AssertionError(f"Row {row_no}: empty {label}")
    # Use a short but meaningful probe. Very short answers such as 「はい。」 are
    # checked as-is; longer cells are matched by their leading visible text.
    probe = text if len(text) <= 36 else text[:36]
    if probe not in source_text:
        raise AssertionError(f"Row {row_no}: {label} text not found in live HTML: {probe!r}")


def main() -> int:
    index_url = read_index_url()
    live_pages = discover_index_pages(index_url)
    if len(live_pages) != EXPECTED_PAGES:
        raise AssertionError(f"Expected {EXPECTED_PAGES} live FAQ pages, got {len(live_pages)}")

    with tempfile.TemporaryDirectory(prefix="moj-isa-faq-e2e-") as tmpdir:
        output = Path(tmpdir) / "moj_isa_faq.xlsx"
        cmd = [
            sys.executable,
            str(SCRIPT),
            "--url-file",
            str(URL_FILE),
            "--output",
            str(output),
            "--sleep",
            "0",
            "--expected-page-count",
            str(EXPECTED_PAGES),
            "--min-total-qa",
            str(MIN_EXPECTED_ROWS),
        ]
        completed = subprocess.run(cmd, cwd=ROOT, text=True, capture_output=True)
        print(completed.stdout, end="")
        if completed.stderr:
            print(completed.stderr, file=sys.stderr, end="")
        if completed.returncode != 0:
            return completed.returncode
        if not output.exists() or output.stat().st_size == 0:
            raise AssertionError("Excel output was not created")

        workbook = load_workbook(output, read_only=True, data_only=True)
        for sheet in ["QA", "Pages", "Summary"]:
            if sheet not in workbook.sheetnames:
                raise AssertionError(f"Missing sheet: {sheet}")

        qa_rows = sheet_records(workbook, "QA")
        page_rows = sheet_records(workbook, "Pages")
        required_headers = {
            "page_no",
            "faq_page_title",
            "category",
            "section",
            "question_no",
            "question",
            "answer",
            "faq_page_url",
            "answer_page_url",
        }
        missing = required_headers.difference(qa_rows[0].keys() if qa_rows else set())
        if missing:
            raise AssertionError(f"Missing QA columns: {sorted(missing)}")
        if len(qa_rows) < MIN_EXPECTED_ROWS:
            raise AssertionError(f"Expected at least {MIN_EXPECTED_ROWS} QA rows, got {len(qa_rows)}")

        workbook_pages = [
            (normalize_space(row["faq_page_title"]), normalize_space(row["faq_page_url"]))
            for row in page_rows
        ]
        if workbook_pages != live_pages:
            raise AssertionError(f"Workbook Pages sheet does not match live index pages: {workbook_pages!r} != {live_pages!r}")

        page_count_from_qa: dict[str, int] = {}
        for row in qa_rows:
            title = normalize_space(row["faq_page_title"])
            page_count_from_qa[title] = page_count_from_qa.get(title, 0) + 1
        for page in page_rows:
            title = normalize_space(page["faq_page_title"])
            expected = int(page["qa_rows"])
            actual = page_count_from_qa.get(title, 0)
            if actual != expected:
                raise AssertionError(f"Pages.qa_rows mismatch for {title}: {expected} != {actual}")

        source_texts = visible_text_cache({normalize_space(row["answer_page_url"]) for row in qa_rows})
        for pos, row in enumerate(qa_rows, start=2):
            answer_url = normalize_space(row["answer_page_url"])
            base_url, _fragment = urldefrag(answer_url)
            source_text = source_texts[base_url]
            assert_text_appears("question", row["question"], source_text, pos)
            assert_text_appears("answer", row["answer"], source_text, pos)

        print(f"E2E OK: {len(qa_rows)} rows, {len(page_rows)} pages; workbook text matched live HTML")
        return 0


if __name__ == "__main__":
    raise SystemExit(main())
